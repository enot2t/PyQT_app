#!pip install PyQt5

import sys
import openpyxl
import datetime
import sqlite3 as sq
from itertools import chain
from PyQt5.QtWidgets import QApplication, QDialog, QWidget, QStatusBar, QPushButton, QLabel, QVBoxLayout, QHBoxLayout, \
    QTextEdit, QInputDialog, QTableWidget, QTableWidgetItem
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QStyledItemDelegate
from PyQt5.QtGui import QColor, QFont


class HighlightDelegate(QStyledItemDelegate):
    def __init__(self, parent=None):
        super().__init__(parent)

    def paint(self, painter, option, index):
        value = index.data()
        if value == 'OK':
            # Подсветить ячейку зеленым фоном, если значение 'OK'
            painter.fillRect(option.rect, QColor(0, 255, 0))
        elif value == 'НЕТ':
            # Подсветить ячейку зеленым фоном, если значение 'НЕТ'
            painter.fillRect(option.rect, QColor(255, 0, 0))
        else:
            # Для других значений используется стандартное отображение
            super().paint(painter, option, index)


class SimpleApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.telega = []

    def initUI(self):
        # кнопки
        self.button = QPushButton('Привязать тележку', self)
        self.button_2 = QPushButton('Отвязать тележку', self)
        self.button_status = QPushButton('Привязанные тележки', self)
        self.button_save_excel = QPushButton('Сохранить в Excel', self)
        self.button_exit = QPushButton('Выход', self)
        self.button.clicked.connect(self.on_button_click)
        self.button_2.clicked.connect(self.on_button_click_2)
        self.button_status.clicked.connect(self.on_button_click_status)
        self.button_exit.clicked.connect(self.on_button_exit)
        self.button_save_excel.clicked.connect(self.on_button_save_excel)

        # соединение с базой данных
        self.conn = sq.connect('my_database.db')
        self.cursor = self.conn.cursor()

        self.cursor.execute('''CREATE TABLE IF NOT EXISTS cold_elements (
                                CELL TEXT PRIMARY KEY,
                                TELEGA TEXT,
                                START_DATE TEXT,
                                END_DATE TEXT)''')
        self.conn.commit()

        # текстовое поле
        self.text_edit = QTextEdit(self)
        self.text_edit.setReadOnly(True)

        # задаем виджет таблицы
        self.table = QTableWidget(self)
        self.table.setColumnCount(5)  # Количество колонок
        self.table.setHorizontalHeaderLabels(["CELL", "TELEGA", "START_DATE", "END_DATE", "STATUS"])
        self.table.horizontalHeader().setStretchLastSection(True)

        # наследование класса с цветовым обозначением
        self.table.setItemDelegate(HighlightDelegate(self))

        # верхний слой кнопок
        button_layout_up = QHBoxLayout()
        button_layout_up.addWidget(self.button)
        button_layout_up.addWidget(self.button_2)

        # нижний слой кнопок
        button_layout_down = QHBoxLayout()
        button_layout_down.addWidget(self.button_status)
        button_layout_down.addWidget(self.button_save_excel)
        button_layout_down.addWidget(self.button_exit)

        self.statusBar = QWidget()
        self.statusBarLayout = QHBoxLayout(self.statusBar)

        # Устанавливаем версию приложения в статус-бар
        version = "Версия 1.01"
        version_label = QLabel(version, self)
        version_label.setAlignment(Qt.AlignRight)
        self.statusBarLayout.addWidget(version_label)

        # основной слой
        main_layout = QVBoxLayout()
        main_layout.addLayout(button_layout_up)
        main_layout.addLayout(button_layout_down)
        main_layout.addWidget(self.text_edit)
        main_layout.addWidget(self.table)
        main_layout.addWidget(self.statusBar)

        self.setLayout(main_layout)

        # размеры приложения
        # self.setGeometry(100, 100, 400, 300)
        self.setWindowTitle('Привязка ХЭ')

        self.resize(400, 800)  # Set the size of the main window
        self.table.setFixedSize(780, 300)
        self.show()

    def on_b1(self):
        self.dlg.close()

    def showdialog(self):
        self.dlg = QDialog()
        self.dlg.resize(150, 150)
        self.b1 = QPushButton("Прочитано", self.dlg)
        self.b1.setAutoDefault(False)
        self.b1.clicked.connect(self.on_b1)

        # self.b1.move(100,100)
        self.dlg.setWindowTitle("ОШИБКА!!!")

        vbox = QVBoxLayout()
        self.label = QLabel("   ОШИБКА!!!")
        self.label_2 = QLabel("Срок охлаждения меньше 48 часов")
        self.label.setFont(QFont('Arial', 30))
        self.label_2.setFont(QFont('Arial', 20))
        vbox.addWidget(self.label)
        vbox.addWidget(self.label_2)
        vbox.addWidget(self.b1)
        self.dlg.setLayout(vbox)
        # self.dlg.setGeometry(100, 100, 400, 300)
        self.dlg.resize(300, 300)
        self.dlg.setWindowModality(Qt.ApplicationModal)
        self.dlg.exec_()

    def status_telega(self, cell, telega):
        dt_now = datetime.datetime.now().replace(microsecond=0).strftime('%Y-%m-%d %H:%M:%S')
        dt_now_test = '2023-11-05 08:00:00'
        if cell != 1:
            self.cursor.execute(f"SELECT STATUS FROM(SELECT *, CASE WHEN END_DATE < '{dt_now_test}' THEN 'OK' ELSE 'НЕТ' \
                    END AS STATUS FROM cold_elements WHERE CELL LIKE '{cell}')")
            rows = self.cursor.fetchall()
            return rows[0][0]
        else:
            self.cursor.execute(f"SELECT STATUS FROM(SELECT *, CASE WHEN END_DATE < '{dt_now_test}' THEN 'OK' ELSE 'НЕТ' \
                    END AS STATUS FROM cold_elements WHERE TELEGA LIKE '{telega}')")
            rows = self.cursor.fetchall()
            return rows[0][0]

    def display_database_data(self):
        # Выполняем SELECT-запрос для получения всех записей
        dt_now = datetime.datetime.now().replace(microsecond=0).strftime('%Y-%m-%d %H:%M:%S')
        dt_now_test = '2023-11-05 08:00:00'
        self.cursor.execute(f"SELECT *, CASE WHEN END_DATE < '{dt_now_test}' THEN 'OK' ELSE 'НЕТ' \
        END AS STATUS FROM cold_elements ORDER BY END_DATE DESC")
        rows = self.cursor.fetchall()

        self.table.setRowCount(0)

        for row in rows:
            self.table.insertRow(self.table.rowCount())
            for i, value in enumerate(row):
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(self.table.rowCount() - 1, i, item)

        # self.table.resizeColumnsToContents()
        column_width = 150  # ширина колонки
        for i in range(self.table.columnCount()):
            self.table.setColumnWidth(i, column_width)

        for row in range(self.table.rowCount()):
            item = self.table.item(row, 2)  # предполагается, что колонка времени
            if item.text() == 'OK':
                item.setData(Qt.DisplayRole, 'OK')

    def closeEvent(self, event):
        # Закрываем соединение с базой данных при завершении приложения
        self.conn.close()
        event.accept()

    def on_button_click(self):
        ok = 'Ok'
        while ok:
            self.telega = []
            text, ok = QInputDialog.getText(self, 'Сканирование ячейки', 'Введите номер ячейки:')
            self.cursor.execute('SELECT CELL FROM cold_elements')
            row = list(chain.from_iterable([list(i) for i in self.cursor.fetchall()]))
            if str(text) not in row:
                if ok and len(text) == 5 and text[:2] == 'MT':
                    self.telega.append(text)
                    self.text_edit.append('Введена ячейка: ' + text)
                    cell, ok = QInputDialog.getText(self, 'Сканирование тележки', 'Введите номер тележки:')
                    self.telega.append(cell)
                    self.text_edit.append('Введена тележка: ' + cell)
                    self.cursor.execute(f'SELECT TELEGA FROM cold_elements')
                    row_cell = list(chain.from_iterable([list(i) for i in self.cursor.fetchall()]))
                    if ok:
                        if str(cell) in row_cell:
                            self.text_edit.append('Ошибка: тележка уже привязана')
                            continue  #
                        else:
                            # output_text = ', '.join(self.telej)
                            first_item = self.telega[0]
                            second_item = self.telega[1]
                            dt = datetime.datetime.now().replace(microsecond=0)
                            dt_end = dt + datetime.timedelta(days=2)
                            self.text_edit.append(f'Привязаны: Ячейка - {first_item}, Тележка -  {second_item}')
                            self.cursor.execute(
                                "INSERT INTO cold_elements ( CELL, TELEGA, START_DATE, END_DATE) VALUES (?, ?, ?, ?)",
                                (first_item, str(second_item), dt, dt_end))
                            self.conn.commit()  # Сохраняем изменения в базе данных

                            # Отображаем весь список из базы данных
                            self.display_database_data()
                            continue
                    else:
                        self.text_edit.append('Отмена операции')
                        self.close()

                else:
                    self.text_edit.append('Неверный ввод ячейки: ' + text)
                    self.text_edit.append('Отмена операции')
                    continue
            else:
                self.cursor.execute(f'SELECT CELL, TELEGA FROM cold_elements WHERE CELL LIKE "{text}"')
                rows = self.cursor.fetchall()
                self.text_edit.append(f'Ячейка занята: тележка {rows[0][1]}')

    def on_button_click_2(self):
        ok = 'Ok'
        while ok:
            text, ok = QInputDialog.getText(self, 'Удаление тележки', 'Введите номер тележки или ячейки:')
            if ok and len(text) == 5 and text[:2] == 'MT':
                chek_st = self.status_telega(text, 1)
                if chek_st == 'OK':
                    self.cursor.execute(f'DELETE FROM cold_elements WHERE CELL like "{text}"')
                    self.conn.commit()
                    self.text_edit.append('Тележка удалена')
                else:
                    self.text_edit.append('Срок охлаждения меньше 48 часов')
                    self.showdialog()
            else:
                self.cursor.execute(f'SELECT TELEGA FROM cold_elements')
                row_cell = list(chain.from_iterable([list(i) for i in self.cursor.fetchall()]))
                if str(text) in row_cell:
                    chek_st = self.status_telega(1, text)
                    if chek_st == 'OK':
                        self.cursor.execute(f'SELECT CELL FROM cold_elements WHERE TELEGA like "{text}"')
                        cells = self.cursor.fetchall()
                        self.cursor.execute(f'DELETE FROM cold_elements WHERE TELEGA like "{text}"')
                        self.conn.commit()
                        self.text_edit.append(f'Тележка изъята из ячейки: {cells[0][0]}')
                    else:
                        self.text_edit.append('Срок охлаждения меньше 48 часов')
                        self.showdialog()
                else:
                    self.text_edit.append('Неверный ввод')
                    self.text_edit.append('Отмена операции')

    def on_button_save_excel(self):
        dt_now = datetime.datetime.now().replace(microsecond=0).strftime('%Y-%m-%d %H:%M:%S')
        self.cursor.execute(f"SELECT *, CASE WHEN END_DATE < '{dt_now}' THEN 'OK' ELSE 'НЕТ' \
        END AS STATUS FROM cold_elements ORDER BY END_DATE DESC")
        rows = self.cursor.fetchall()

        # Создаем новую книгу Excel и активный лист
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Заголовки столбцов
        columns = ['ЯЧЕЙКА', 'ТЕЛЕЖКА', 'ВРЕМЯ_ПРИВЯЗКИ', 'ОКОНЧАНИЕ_ОХЛАЖД.', 'СТАТУС']
        for col_num, column_title in enumerate(columns, 1):
            sheet.cell(row=1, column=col_num, value=column_title)

        # Записываем данные
        for row_num, row_data in enumerate(rows, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col_num, value=cell_value)

        # Сохраняем книгу
        dt = datetime.datetime.now().date()
        file_path = f'Тележки-{str(dt)}.xlsx'
        workbook.save(file_path)

        self.text_edit.append('Данные выгружены в Excel файл: ' + file_path)

    def on_button_click_status(self):
        self.display_database_data()

    def on_button_exit(self):
        # Close the database connection and exit the application
        self.conn.close()
        sys.exit()


def on_about_to_quit():
    print("Приложение завершается")


def main():
    app = QApplication(sys.argv)
    ex = SimpleApp()

    app.aboutToQuit.connect(on_about_to_quit)

    sys.exit(app.exec_())


if __name__ == '__main__':
    main()